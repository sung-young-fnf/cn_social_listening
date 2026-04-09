"""MST_PROFILE 마스터 등록용 엑셀 생성 스크립트

기존 마스터(CHN_MKT.MST_PROFILE.xlsx)에 없는 계정만 추출하여
담당자에게 전달할 엑셀 파일을 생성합니다.

사용법:
    python generate_mst_profile.py
"""
import os
import json
import openpyxl
from datetime import datetime


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_XLSX = os.path.join(BASE_DIR, "CHN_MKT.MST_PROFILE.xlsx")

# 샤오홍슈 크롤링 결과 폴더들
XHS_DIRS = [
    os.path.join(BASE_DIR, "MediaCrawler", "output", "red-weekly-260313"),
    os.path.join(BASE_DIR, "MediaCrawler", "output", "red-weekly-260316"),
]


def parse_chinese_number(text):
    """한자 숫자 표기를 정수로 변환"""
    if not text or text == "":
        return 0
    text = str(text).replace("+", "").replace(",", "").strip()
    try:
        if "万" in text:
            return int(float(text.replace("万", "")) * 10000)
        elif "亿" in text:
            return int(float(text.replace("亿", "")) * 100000000)
        else:
            return int(float(text))
    except (ValueError, TypeError):
        return 0


def load_existing_profiles():
    """기존 마스터 엑셀에서 등록된 계정 ID 추출"""
    wb = openpyxl.load_workbook(MASTER_XLSX)
    ws = wb.active

    existing = {"DOUYIN": set(), "RED": set()}
    for row in range(4, ws.max_row + 1):
        pid = ws.cell(row=row, column=3).value
        platform = ws.cell(row=row, column=4).value
        if pid and platform in existing:
            existing[platform].add(str(pid))
    return existing


def collect_xhs_accounts():
    """샤오홍슈 크롤링 결과에서 계정 정보 수집 (중복 제거)"""
    accounts = {}
    for xhs_dir in XHS_DIRS:
        if not os.path.isdir(xhs_dir):
            continue
        for folder in os.listdir(xhs_dir):
            cp = os.path.join(xhs_dir, folder, "creator.json")
            if not os.path.isfile(cp):
                continue
            with open(cp, "r", encoding="utf-8") as f:
                c = json.load(f)
            uid = c.get("user_id", "")
            if uid and uid not in accounts:
                accounts[uid] = c
    return accounts


def generate_excel(new_accounts, platform):
    """MST_PROFILE 양식 엑셀 생성"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MST_PROFILE"

    # 헤더 (기존 양식과 동일)
    headers = [
        "PROFILE_ID", "PLATFORM", "PROFILE_TYPE", "PROFILE_NM",
        "FOLLOWER_CNT", "TOTAL_LIKE_CNT", "PROFILE_IMG",
        "PROFILE_URL", "CREATE_DTTM", "UPDATE_DTTM",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S.000000")

    for i, (uid, data) in enumerate(sorted(new_accounts.items(), key=lambda x: x[1].get("nickname", "")), 2):
        if platform == "RED":
            nickname = data.get("nickname", "")
            fans = parse_chinese_number(data.get("fans", 0))
            interaction = parse_chinese_number(data.get("interaction", 0))
            img_path = f"xiaohongshu/account/image/{uid}/{uid}.png"
            profile_url = ""
        elif platform == "DOUYIN":
            nickname = data.get("nickname", "")
            fans = data.get("followerCount", 0)
            interaction = data.get("totalFavorited", 0)
            img_path = data.get("profile_img", "")
            profile_url = data.get("profileUrl", "")

        ws.cell(row=i, column=1, value=uid)
        ws.cell(row=i, column=2, value=platform)
        ws.cell(row=i, column=3, value="")  # PROFILE_TYPE — 수동 분류 필요
        ws.cell(row=i, column=4, value=nickname)
        ws.cell(row=i, column=5, value=fans)
        ws.cell(row=i, column=6, value=interaction)
        ws.cell(row=i, column=7, value=img_path)
        ws.cell(row=i, column=8, value=profile_url)
        ws.cell(row=i, column=9, value=now_str)
        ws.cell(row=i, column=10, value=now_str)

    # 열 너비 조정
    col_widths = [20, 10, 15, 25, 15, 15, 50, 50, 30, 30]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    return wb


def main():
    print("기존 마스터 로드 중...")
    existing = load_existing_profiles()
    print(f"  기존 DOUYIN: {len(existing['DOUYIN'])}개, RED: {len(existing['RED'])}개\n")

    # 샤오홍슈 신규 계정
    print("샤오홍슈 크롤링 결과 수집 중...")
    xhs_accounts = collect_xhs_accounts()
    new_xhs = {uid: data for uid, data in xhs_accounts.items() if uid not in existing["RED"]}
    print(f"  전체: {len(xhs_accounts)}개, 신규: {len(new_xhs)}개\n")

    if new_xhs:
        wb = generate_excel(new_xhs, "RED")
        output_path = os.path.join(BASE_DIR, "MST_PROFILE_NEW_RED.xlsx")
        wb.save(output_path)
        print(f"엑셀 생성 완료: {output_path}")
        print(f"  → {len(new_xhs)}개 계정")
        print(f"  → PROFILE_TYPE 컬럼은 비어있습니다. 수동으로 분류해주세요.")
        print(f"     (인플루언서 / 셀럽 / 자사 / 경쟁사 / 메가페이지)")
    else:
        print("신규 등록할 샤오홍슈 계정이 없습니다.")


if __name__ == "__main__":
    main()
